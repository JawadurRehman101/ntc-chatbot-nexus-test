import io
import datetime
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


class VDSProformaPDF(FPDF):
    def header(self):
        self.set_font("Helvetica", "B", 18)
        self.set_text_color(0, 100, 60)
        self.cell(0, 10, "National Telecommunication Corporation", align="C", new_x="LMARGIN", new_y="NEXT")
        self.set_font("Helvetica", "", 11)
        self.set_text_color(80, 80, 80)
        self.cell(0, 7, "IaaS VDS (Virtual Dedicated Server) - Service Proforma", align="C", new_x="LMARGIN", new_y="NEXT")
        self.set_draw_color(0, 100, 60)
        self.set_line_width(0.8)
        self.line(10, self.get_y() + 2, 200, self.get_y() + 2)
        self.ln(8)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 10, f"Page {self.page_no()}/{{nb}} | Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}", align="C")

    def section_title(self, title):
        self.set_font("Helvetica", "B", 13)
        self.set_fill_color(0, 100, 60)
        self.set_text_color(255, 255, 255)
        self.cell(0, 9, f"  {title}", fill=True, new_x="LMARGIN", new_y="NEXT")
        self.set_text_color(0, 0, 0)
        self.ln(3)

    def add_field_row(self, label, value):
        self.set_font("Helvetica", "B", 10)
        self.cell(70, 7, label, border=1)
        self.set_font("Helvetica", "", 10)
        self.cell(0, 7, str(value), border=1, new_x="LMARGIN", new_y="NEXT")


def generate_vds_pdf(form_data, user_name, user_email):
    from config import BILLING_POC_FIELDS, TECHNICAL_POC_FIELDS, SERVER_DETAIL_FIELDS, GENERAL_DETAIL_FIELDS

    pdf = VDSProformaPDF()
    pdf.alias_nb_pages()
    pdf.add_page()

    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, f"Applicant: {user_name}  |  Email: {user_email}  |  Date: {datetime.datetime.now().strftime('%Y-%m-%d')}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    # Section 1
    pdf.section_title("Section 1: Billing POC Details")
    billing = form_data.get("billing_poc", {})
    for key, label in BILLING_POC_FIELDS:
        pdf.add_field_row(label, billing.get(key, ""))
    pdf.ln(5)

    # Section 2
    pdf.section_title("Section 2: Technical POC Details")
    technical = form_data.get("technical_poc", {})
    for key, label in TECHNICAL_POC_FIELDS:
        pdf.add_field_row(label, technical.get(key, ""))
    pdf.ln(5)

    # Section 3
    pdf.section_title("Section 3: Server Details")
    servers = form_data.get("servers", [])
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 7, f"Total Servers: {form_data.get('server_count', 0)}  |  Configuration: {'Identical' if form_data.get('servers_identical') else 'Unique'}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)
    for i, srv in enumerate(servers, 1):
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(0, 7, f"Server {i}", new_x="LMARGIN", new_y="NEXT")
        for key, label in SERVER_DETAIL_FIELDS:
            pdf.add_field_row(label, srv.get(key, ""))
        pdf.ln(2)
    pdf.ln(3)

    # Section 4
    pdf.section_title("Section 4: General Details")
    general = form_data.get("general", {})
    for key, label in GENERAL_DETAIL_FIELDS:
        pdf.add_field_row(label, general.get(key, ""))

    return pdf.output()


def generate_vds_excel(form_data, user_name, user_email):
    from config import BILLING_POC_FIELDS, TECHNICAL_POC_FIELDS, SERVER_DETAIL_FIELDS, GENERAL_DETAIL_FIELDS

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VDS Service Proforma"

    header_font = Font(name="Arial", size=14, bold=True, color="006432")
    section_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    section_fill = PatternFill(start_color="006432", end_color="006432", fill_type="solid")
    label_font = Font(name="Arial", size=10, bold=True)
    value_font = Font(name="Arial", size=10)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    row = 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "National Telecommunication Corporation - IaaS VDS Proforma"
    ws[f"A{row}"].font = header_font
    row += 1
    ws[f"A{row}"] = f"Applicant: {user_name}"
    ws[f"D{row}"] = f"Email: {user_email}"
    ws[f"G{row}"] = f"Date: {datetime.datetime.now().strftime('%Y-%m-%d')}"
    row += 2

    def write_section(title, fields, data):
        nonlocal row
        ws.merge_cells(f"A{row}:H{row}")
        cell = ws[f"A{row}"]
        cell.value = title
        cell.font = section_font
        cell.fill = section_fill
        row += 1
        for key, label in fields:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = label_font
            ws[f"A{row}"].border = thin_border
            ws.merge_cells(f"B{row}:H{row}")
            ws[f"B{row}"] = str(data.get(key, ""))
            ws[f"B{row}"].font = value_font
            ws[f"B{row}"].border = thin_border
            row += 1
        row += 1

    write_section("Section 1: Billing POC Details", BILLING_POC_FIELDS, form_data.get("billing_poc", {}))
    write_section("Section 2: Technical POC Details", TECHNICAL_POC_FIELDS, form_data.get("technical_poc", {}))

    # Server section
    ws.merge_cells(f"A{row}:H{row}")
    cell = ws[f"A{row}"]
    cell.value = "Section 3: Server Details"
    cell.font = section_font
    cell.fill = section_fill
    row += 1
    ws[f"A{row}"] = f"Total Servers: {form_data.get('server_count', 0)}"
    ws[f"D{row}"] = f"Config: {'Identical' if form_data.get('servers_identical') else 'Unique'}"
    row += 1

    for col_idx, (_, label) in enumerate(SERVER_DETAIL_FIELDS):
        cell = ws.cell(row=row, column=col_idx + 1, value=label)
        cell.font = label_font
        cell.fill = PatternFill(start_color="DCE9DF", end_color="DCE9DF", fill_type="solid")
        cell.border = thin_border
    row += 1
    for server in form_data.get("servers", []):
        for col_idx, (key, _) in enumerate(SERVER_DETAIL_FIELDS):
            cell = ws.cell(row=row, column=col_idx + 1, value=str(server.get(key, "")))
            cell.font = value_font
            cell.border = thin_border
        row += 1
    row += 1

    write_section("Section 4: General Details", GENERAL_DETAIL_FIELDS, form_data.get("general", {}))

    from openpyxl.utils import get_column_letter
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class ColocationRequestPDF(FPDF):
    def header(self):
        self.set_font("Helvetica", "B", 18)
        self.set_text_color(0, 100, 60)
        self.cell(0, 10, "National Telecommunication Corporation", align="C", new_x="LMARGIN", new_y="NEXT")
        self.set_font("Helvetica", "", 11)
        self.set_text_color(80, 80, 80)
        self.cell(0, 7, "Data Center Colocation Services Request Form", align="C", new_x="LMARGIN", new_y="NEXT")
        self.set_draw_color(0, 100, 60)
        self.set_line_width(0.8)
        self.line(10, self.get_y() + 2, 200, self.get_y() + 2)
        self.ln(8)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 10, f"Page {self.page_no()}/{{nb}} | Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}", align="C")

    def section_title(self, title):
        self.set_font("Helvetica", "B", 13)
        self.set_fill_color(0, 100, 60)
        self.set_text_color(255, 255, 255)
        self.cell(0, 9, f"  {title}", fill=True, new_x="LMARGIN", new_y="NEXT")
        self.set_text_color(0, 0, 0)
        self.ln(3)

    def add_field_row(self, label, value):
        self.set_font("Helvetica", "B", 10)
        self.cell(85, 7, label, border=1)
        self.set_font("Helvetica", "", 10)
        self.cell(0, 7, str(value), border=1, new_x="LMARGIN", new_y="NEXT")


def generate_colocation_pdf(form_data, user_name, user_email):
    from config import COLOCATION_BILLING_FIELDS, COLOCATION_REQUIREMENT_FIELDS

    pdf = ColocationRequestPDF()
    pdf.alias_nb_pages()
    pdf.add_page()

    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, f"Applicant: {user_name}  |  Email: {user_email}  |  Date: {datetime.datetime.now().strftime('%Y-%m-%d')}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    # Section 1: Billing Details
    pdf.section_title("Billing Details")
    billing = form_data.get("billing", {})
    for key, label in COLOCATION_BILLING_FIELDS:
        pdf.add_field_row(label, billing.get(key, ""))
    pdf.ln(5)

    # Section 2: Colocation Requirements
    pdf.section_title("Colocation Requirements")
    reqs = form_data.get("requirements", {})
    
    # Table headers
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_fill_color(220, 233, 223)
    pdf.cell(15, 8, "S. No.", border=1, fill=True, align="C")
    pdf.cell(100, 8, "Item Description", border=1, fill=True)
    pdf.cell(30, 8, "Unit", border=1, fill=True, align="C")
    pdf.cell(0, 8, "Required (Qty)", border=1, fill=True, new_x="LMARGIN", new_y="NEXT", align="C")
    
    pdf.set_font("Helvetica", "", 9)
    units = {
        "rack_space_42u": "42 Rack",
        "rack_space_ru": "RU",
        "power_kwh": "KWh",
        "internet_uplink_mbps": "Mbps",
        "network_security": "Mbps",
        "ssl_vpns": "",
        "endpoint_security": "",
        "ssl_cert_type": "",
        "domain_registration": "",
        "public_ip": "",
        "other_requirements": ""
    }
    
    for idx, (key, label) in enumerate(COLOCATION_REQUIREMENT_FIELDS, 1):
        unit = units.get(key, "")
        val = reqs.get(key, "")
        pdf.cell(15, 8, str(idx), border=1, align="C")
        pdf.cell(100, 8, label[:55], border=1)
        pdf.cell(30, 8, unit, border=1, align="C")
        pdf.cell(0, 8, str(val), border=1, new_x="LMARGIN", new_y="NEXT")
        
    return pdf.output()


def generate_colocation_excel(form_data, user_name, user_email):
    from config import COLOCATION_BILLING_FIELDS, COLOCATION_REQUIREMENT_FIELDS

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Colocation Request"

    header_font = Font(name="Arial", size=14, bold=True, color="006432")
    section_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    section_fill = PatternFill(start_color="006432", end_color="006432", fill_type="solid")
    label_font = Font(name="Arial", size=10, bold=True)
    value_font = Font(name="Arial", size=10)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    # Title Row
    ws.merge_cells("A1:D1")
    ws["A1"] = "Data Center Colocation Services Request Form"
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color="006432")
    ws["A1"].alignment = Alignment(horizontal="center")
    
    ws["A2"] = f"Applicant: {user_name}  |  Email: {user_email}  |  Date: {datetime.datetime.now().strftime('%Y-%m-%d')}"
    ws["A2"].font = value_font
    ws.merge_cells("A2:D2")
    ws["A2"].alignment = Alignment(horizontal="center")
    
    row = 4
    
    # Section: Billing Details
    ws.merge_cells(f"A{row}:D{row}")
    cell = ws[f"A{row}"]
    cell.value = "Billing Details"
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = Alignment(horizontal="center")
    row += 1
    
    billing = form_data.get("billing", {})
    for key, label in COLOCATION_BILLING_FIELDS:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = label_font
        ws[f"A{row}"].border = thin_border
        
        ws.merge_cells(f"B{row}:D{row}")
        ws[f"B{row}"] = str(billing.get(key, ""))
        ws[f"B{row}"].font = value_font
        
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = thin_border
            
        row += 1
        
    row += 2
    
    # Section: Requirements Table
    headers = ["S. No.", "Item Description", "Unit", "Required (Qty)"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = label_font
        cell.fill = PatternFill(start_color="DCE9DF", end_color="DCE9DF", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        
    row += 1
    
    units = {
        "rack_space_42u": "42 Rack",
        "rack_space_ru": "RU",
        "power_kwh": "KWh",
        "internet_uplink_mbps": "Mbps",
        "network_security": "Mbps",
        "ssl_vpns": "",
        "endpoint_security": "",
        "ssl_cert_type": "",
        "domain_registration": "",
        "public_ip": "",
        "other_requirements": ""
    }
    
    reqs = form_data.get("requirements", {})
    for idx, (key, label) in enumerate(COLOCATION_REQUIREMENT_FIELDS, 1):
        unit = units.get(key, "")
        val = reqs.get(key, "")
        
        ws.cell(row=row, column=1, value=idx).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=label)
        ws.cell(row=row, column=3, value=unit).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=str(val))
        
        for c in range(1, 5):
            cell = ws.cell(row=row, column=c)
            cell.font = value_font
            cell.border = thin_border
            
        row += 1
        
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, 5):
        max_len = 0
        for r in range(1, row):
            val = str(ws.cell(row=r, column=col_idx).value or "")
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)
        
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

